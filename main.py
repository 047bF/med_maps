# -*- coding: utf-8 -*-
from optparse import OptionParser
import openpyxl
import re

import psycopg2 as psycopg2
from psycopg2 import sql

import requests
from bs4 import BeautifulSoup

yandex = '83702ac8-dc4d-4ac7-9dd4-37f625766a9c'

# Press ⌃R to execute it or replace it with your code.
# Press Double ⇧ to search everywhere for classes, files, tool windows, actions, and settings.

"""
добавить:
    пропуск страниц в параметрах (как в начале так и в конце, перечислением)
    пропуск лишних строк в параметрах (как в начале так и в конце, перечислением)
    работа в цикле по странице
"""

conn = psycopg2.connect("dbname=postgres user=postgres password=secret")
cursor = conn.cursor()

insert_query = sql.SQL("insert into newtable (address, point) values (%s, %s)")
select_query = sql.SQL("select point from newtable where address = %s")


def convertTuple(tup):
    converted_tup = ''.join(tup)
    return converted_tup


def point_process(address):
    cursor.execute(select_query, (address, ))
    record = cursor.fetchall()
    if len(record) == 0:
        response = requests.get('https://geocode-maps.yandex.ru/1.x/?apikey=' + yandex + '&geocode=' + address)
        if response.status_code == 403:
            return 403
        soup = BeautifulSoup(response.text, "html.parser")
        if soup.find('ymaps').find_all('pos'):
            soup_point = soup.find('ymaps').find('pos').text
            cursor.execute(insert_query, (address, soup_point))
            conn.commit()
        else:
            return 404
        record = soup_point
    else:
        record = convertTuple(record[0])
    return record


def parse_sheets(wb_obj):
    new_val = 0
    sheets = wb_obj.sheetnames
    for sheet in sheets:
        cur_sheet = wb_obj[sheet]
        if sheet == 'Общая информация':
            continue
        else:
            parse_map(cur_sheet)
            new_val += 1
            print('End page ' + str(new_val))


def parse_arg():
    parser = OptionParser()
    parser.add_option("-f", "--file", dest="filename",
                      help="write report to FILE", metavar="FILE")
    parser.add_option("-q", "--quiet",
                      action="store_false", dest="verbose", default=True,
                      help="don't print status messages to stdout")
    (options, args) = parser.parse_args()
    return options


def parse_page(page):
    #print(page)
    # sheet = wb_obj["1.1.Бизнес"]
    # print(wb_obj.sheetnames)
    # print(page["A30"].value)
    for row in page.iter_rows(max_row=6):
        for cell in row:
            print(cell.value, end=" ")
        #print()


def get_point_type(cell):
    point_type = None
    if cell == 'Поликлиника':
        point_type = 'pm2gnm'
    elif cell == 'Плановая и экстренная стационарная помощь':
        point_type = 'pm2blm'
    elif cell == 'Скорая медицинская помощь':
        point_type = 'pm2dbm'
    elif cell in ['Стоматология со свободным выбором', 'Стоматология с прикреплением']:
        point_type = 'pm2grm'
    return point_type


def parse_map(sheet):
    sheet_row_num = 0
    full_row_on_sheet = 0
    point_count = 0
    point = ''
    point_type = 'pm2grm'
    for row in sheet.rows:
        sheet_row_num += 1
        #print(sheet_row_num)
        if sheet_row_num < 2:
            continue
        if row[0].value != None:
            #elif sheet_row_num > 181:
            #    break
            cell = row[0].value
            clean_cell = make_clean_cell(cell)
            cell_row_num = 0
            org_name = ''
            for row in clean_cell:
                full_row_on_sheet += 1
                cell_row_num += 1
                if cell_row_num == 1:
                    if get_point_type(row) is None:
                        org_name = row
                    else:
                        point_type = get_point_type(row)
                else:
                    point_processed = point_process(row)
                    if point_processed == 403 or point_processed == 404:
                        print(row)
                        print(str(sheet_row_num) + ' 403 error')
                        break
                    elif point_processed == 404:
                        print(str(sheet_row_num) + ' 404 error')
                        break
                    if (row.find('Санкт-Петербург') > 1 or row.find('Санкт Петербург') > 1) and row.find('Сестрорецк') < 1 and row.find('Песочный') < 1 and row.find('Колпино') < 1 and row.find('Кронштадт') < 1 and row.find('Пушкин') < 1 and row.find('Павловск') < 1 and row.find('Петергоф ') < 1:
                        point += point_processed.replace(' ', ',') + ',' + point_type + '~'
                        point_count += 1
                    if point_count >= 100:
                        print('https://static-maps.yandex.ru/1.x/?ll=30.361954,59.950406&z=11&l=map&pt=' + point[:-1])
                        map_point = requests.get('https://static-maps.yandex.ru/1.x/?l=map&pt=' + point[:-1])
                        point_count = 1
                        point = ''
    if point != '':
        print('https://static-maps.yandex.ru/1.x/?ll=30.361954,59.950406&z=11&l=map&pt=' + point[:-1])
        print(row)
        map_point = requests.get('https://static-maps.yandex.ru/1.x/?l=map&pt=' + point[:-1])



def make_clean_cell(line):
    if line is not None:
        return re.sub('\s{2,}', '\n', line).splitlines()


def main():
    options = parse_arg()
    parse_sheets(openpyxl.load_workbook(options.filename))


# def print_hi(name):
#     # Use a breakpoint in the code line below to debug your script.
#     print(f'Hi, {name}')  # Press ⌘F8 to toggle the breakpoint.


# Press the green button in the gutter to run the script.
if __name__ == '__main__':
    main()

# See PyCharm help at https://www.jetbrains.com/help/pycharm/
